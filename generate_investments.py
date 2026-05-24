import os
import glob
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INVESTMENTS_DIR = os.path.join(BASE_DIR, "Investments")
HISTORY_FILE = os.path.join(BASE_DIR, "holdings_history.csv")
OUTPUT_FILE = os.path.join(BASE_DIR, "investments.xlsx")

MANAGED_SHEETS = ["Portfolio Summary", "Holdings", "Transactions", "Dividends", "Account History"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")

INCOME_ACTIONS = {"DIV", "WHTX02", "NRT", "TXPDDV"}
SKIP_ACTIONS = {"SPLIT", "CXLSPL"}


def parse_account_folder(folder_name):
    parts = folder_name.split("_")
    return parts[0], parts[1], parts[2]  # account_type, currency, account_id


def parse_filename_date(filename):
    m = re.search(r"(\d{1,2}-\w{3}-\d{4})", filename)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d-%b-%Y")
        except ValueError:
            pass
    return datetime.min


def find_latest_file(folder_path, file_type):
    files = glob.glob(os.path.join(folder_path, f"*-{file_type}-*.csv"))
    if not files:
        return None
    return max(files, key=lambda f: parse_filename_date(os.path.basename(f)))


def parse_date(d):
    d = str(d).strip()
    for fmt in ("%d-%b-%y", "%d %b %Y"):
        try:
            return datetime.strptime(d, fmt).date()
        except ValueError:
            continue
    return None


def parse_activity(path, account_type, currency, account_id):
    # Rows 0-2: metadata, row 3: header
    df = pd.read_csv(path, skiprows=3, header=0, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.dropna(subset=["Trade Date"])
    df = df[df["Trade Date"].str.strip() != ""]

    df["Date"] = df["Trade Date"].apply(parse_date)
    df = df.dropna(subset=["Date"])

    for col in ["Quantity", "Price", "Commission", "Net Amount"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["Account Type"] = account_type
    df["Currency"] = currency
    df["Account ID"] = account_id

    keep = ["Date", "Account Type", "Account ID", "Currency",
            "Action", "Description", "Quantity", "Price", "Commission",
            "Net Amount", "Security Type"]
    return df[[c for c in keep if c in df.columns]]


def parse_holdings(path, account_type, currency, account_id):
    meta = {"account_type": account_type, "currency": currency, "account_id": account_id}
    header_row_num = None

    with open(path, "r", encoding="utf-8-sig") as f:
        for line_num, line in enumerate(f):
            parts = [p.strip() for p in line.split(",")]
            key = parts[0]
            val = parts[1] if len(parts) > 1 else ""
            if key == "As of Date":
                try:
                    meta["as_of_date"] = pd.to_datetime(val).date()
                except Exception:
                    meta["as_of_date"] = None
            elif key == "Cash":
                meta["cash"] = float(val) if val else 0.0
            elif key == "Investments":
                meta["investments"] = float(val) if val else 0.0
            elif key == "Total Value":
                meta["total_value"] = float(val) if val else 0.0
            elif key == "Symbol":
                header_row_num = line_num
                break

    if header_row_num is None:
        return meta, pd.DataFrame()

    holdings = pd.read_csv(path, skiprows=header_row_num, header=0, dtype=str)
    holdings.columns = [c.strip() for c in holdings.columns]
    holdings = holdings.dropna(subset=["Symbol"])
    holdings = holdings[holdings["Symbol"].str.strip() != ""]

    for col in ["Quantity", "Average Cost", "Price", "Book Cost",
                "Market Value", "Unrealized $", "Unrealized %"]:
        if col in holdings.columns:
            holdings[col] = pd.to_numeric(holdings[col], errors="coerce")

    holdings["Account Type"] = account_type
    holdings["Currency"] = currency
    holdings["Account ID"] = account_id

    keep = ["Account Type", "Account ID", "Currency", "Symbol", "Description",
            "Quantity", "Average Cost", "Price", "Book Cost",
            "Market Value", "Unrealized $", "Unrealized %"]
    return meta, holdings[[c for c in keep if c in holdings.columns]]


def load_history():
    if os.path.exists(HISTORY_FILE):
        try:
            df = pd.read_csv(HISTORY_FILE)
            df["Date"] = pd.to_datetime(df["Date"]).dt.date
            return df
        except Exception:
            pass
    return pd.DataFrame(columns=["Date", "Account Type", "Account ID", "Currency",
                                  "Cash", "Investments", "Total Value"])


def update_history(history_df, snapshots):
    new_rows = []
    for snap in snapshots:
        date = snap.get("as_of_date")
        account_id = snap.get("account_id")
        if date is None:
            continue
        exists = (
            not history_df.empty and
            ((history_df["Date"] == date) & (history_df["Account ID"] == account_id)).any()
        )
        if not exists:
            new_rows.append({
                "Date": date,
                "Account Type": snap["account_type"],
                "Account ID": snap["account_id"],
                "Currency": snap["currency"],
                "Cash": snap.get("cash", 0),
                "Investments": snap.get("investments", 0),
                "Total Value": snap.get("total_value", 0),
            })
    if new_rows:
        history_df = pd.concat([history_df, pd.DataFrame(new_rows)], ignore_index=True)
        history_df.to_csv(HISTORY_FILE, index=False)
        print(f"  Added {len(new_rows)} new snapshot(s) to history.")
    else:
        print("  History already up to date for all accounts.")
    return history_df


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 55)


def color_column(ws, col_letter, last_row):
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        f"{col_letter}2:{col_letter}{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    )


def write_summary_sheet(wb, snapshots):
    ws = wb.create_sheet("Portfolio Summary")
    headers = ["Account Type", "Account ID", "Currency", "Cash", "Investments", "Total Value"]
    ws.append(headers)
    style_header(ws, len(headers))

    rows = sorted(snapshots, key=lambda s: (s["account_type"], s["account_id"]))
    for i, snap in enumerate(rows, start=2):
        ws.append([snap["account_type"], snap["account_id"], snap["currency"],
                   snap.get("cash", 0), snap.get("investments", 0), snap.get("total_value", 0)])
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(i, col)
            cell.fill = fill
            if col >= 4:
                cell.number_format = "#,##0.00"

    total_row = len(rows) + 2
    ws.cell(total_row, 1).value = "TOTAL"
    ws.cell(total_row, 1).font = Font(bold=True)
    ws.cell(total_row, 3).value = "(mixed currencies)"
    ws.cell(total_row, 3).font = Font(italic=True, color="888888")
    for col_idx, key in enumerate(["cash", "investments", "total_value"], start=4):
        cell = ws.cell(total_row, col_idx)
        cell.value = sum(s.get(key, 0) for s in snapshots)
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL

    auto_width(ws)


def write_holdings_sheet(wb, df):
    ws = wb.create_sheet("Holdings")
    if df.empty:
        ws.append(["No holdings data."])
        return

    df = df.sort_values(["Account Type", "Account ID", "Symbol"])
    headers = list(df.columns)
    ws.append(headers)
    style_header(ws, len(headers))

    currency_cols = {"Book Cost", "Market Value", "Unrealized $", "Average Cost", "Price"}

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(i, col_idx)
            cell.fill = fill
            if col_name in currency_cols:
                cell.number_format = "#,##0.00"
            elif col_name == "Unrealized %":
                cell.number_format = '0.00"%"'

    if "Unrealized $" in headers:
        col_letter = get_column_letter(headers.index("Unrealized $") + 1)
        color_column(ws, col_letter, len(df) + 1)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    auto_width(ws)


def write_transactions_sheet(wb, df):
    ws = wb.create_sheet("Transactions")
    if df.empty:
        ws.append(["No transaction data."])
        return

    df = df[~df["Action"].isin(SKIP_ACTIONS)].sort_values("Date", ascending=False)
    headers = list(df.columns)
    ws.append(headers)
    style_header(ws, len(headers))

    amount_col = headers.index("Net Amount") + 1 if "Net Amount" in headers else None

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).fill = fill
        if amount_col:
            ws.cell(i, amount_col).number_format = "#,##0.00"

    if amount_col and len(df) > 0:
        color_column(ws, get_column_letter(amount_col), len(df) + 1)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    auto_width(ws)


def write_dividends_sheet(wb, df):
    ws = wb.create_sheet("Dividends")
    div_df = df[df["Action"].isin(INCOME_ACTIONS)].copy() if not df.empty else pd.DataFrame()
    if div_df.empty:
        ws.append(["No dividend data."])
        return

    div_df = div_df.sort_values("Date", ascending=False)
    keep = [c for c in ["Date", "Account Type", "Account ID", "Currency",
                         "Description", "Action", "Net Amount"] if c in div_df.columns]
    div_df = div_df[keep]
    headers = list(div_df.columns)
    ws.append(headers)
    style_header(ws, len(headers))

    amount_col = headers.index("Net Amount") + 1 if "Net Amount" in headers else None

    for i, row in enumerate(div_df.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).fill = fill
        if amount_col:
            ws.cell(i, amount_col).number_format = "#,##0.00"

    if amount_col and len(div_df) > 0:
        color_column(ws, get_column_letter(amount_col), len(div_df) + 1)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    auto_width(ws)


def write_history_sheet(wb, history_df):
    ws = wb.create_sheet("Account History")
    if history_df.empty:
        ws.append(["No history yet — re-run after adding new holdings files."])
        return

    history_df = history_df.sort_values(
        ["Date", "Account Type", "Account ID"], ascending=[False, True, True]
    )
    headers = list(history_df.columns)
    ws.append(headers)
    style_header(ws, len(headers))

    for i, row in enumerate(history_df.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(i, col_idx)
            cell.fill = fill
            if col_name in {"Cash", "Investments", "Total Value"}:
                cell.number_format = "#,##0.00"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    auto_width(ws)


def main():
    print("Scanning investment accounts...")
    all_activity = []
    all_holdings = []
    snapshots = []

    for folder_name in sorted(os.listdir(INVESTMENTS_DIR)):
        folder_path = os.path.join(INVESTMENTS_DIR, folder_name)
        if not os.path.isdir(folder_path):
            continue
        parts = folder_name.split("_")
        if len(parts) < 3:
            continue
        account_type, currency, account_id = parts[0], parts[1], parts[2]
        print(f"  {account_type} {currency} ({account_id})")

        activity_paths = glob.glob(os.path.join(folder_path, "*-activity-*.csv"))
        if activity_paths:
            for activity_path in activity_paths:
                try:
                    df = parse_activity(activity_path, account_type, currency, account_id)
                    all_activity.append(df)
                except Exception as e:
                    print(f"    Warning: {os.path.basename(activity_path)} failed: {e}")
        else:
            print(f"    No activity file found")

        holdings_path = find_latest_file(folder_path, "holdings")
        if holdings_path:
            try:
                meta, holdings_df = parse_holdings(holdings_path, account_type, currency, account_id)
                snapshots.append(meta)
                if not holdings_df.empty:
                    all_holdings.append(holdings_df)
                    print(f"    Holdings: {len(holdings_df)} positions, total ${meta.get('total_value', 0):,.2f} {currency}")
                else:
                    print(f"    Holdings: empty account")
            except Exception as e:
                print(f"    Warning: holdings parse failed: {e}")

    print("\nUpdating history...")
    history_df = load_history()
    history_df = update_history(history_df, snapshots)

    activity_df = (
        pd.concat(all_activity, ignore_index=True)
        .drop_duplicates(subset=["Date", "Account ID", "Description", "Action", "Quantity", "Net Amount"])
        .reset_index(drop=True)
    ) if all_activity else pd.DataFrame()
    holdings_df = pd.concat(all_holdings, ignore_index=True) if all_holdings else pd.DataFrame()

    print(f"\nWriting {OUTPUT_FILE}...")
    if os.path.exists(OUTPUT_FILE):
        wb = load_workbook(OUTPUT_FILE)
        for name in MANAGED_SHEETS:
            if name in wb.sheetnames:
                del wb[name]
        print(f"  Preserved custom sheets: {[s for s in wb.sheetnames]}" if wb.sheetnames else "  No custom sheets found.")
    else:
        wb = Workbook()
        wb.remove(wb.active)
    write_summary_sheet(wb, snapshots)
    write_holdings_sheet(wb, holdings_df)
    write_transactions_sheet(wb, activity_df)
    write_dividends_sheet(wb, activity_df)
    write_history_sheet(wb, history_df)
    wb.save(OUTPUT_FILE)
    print("Done! Open investments.xlsx to view your portfolio.")


if __name__ == "__main__":
    main()
