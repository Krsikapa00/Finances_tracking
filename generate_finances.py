import os
import glob
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATEMENTS_DIR = os.path.join(BASE_DIR, "Statements")
OUTPUT_FILE = os.path.join(BASE_DIR, "finances.xlsx")
RULES_BACKUP = os.path.join(BASE_DIR, "category_rules_backup.json")

DEFAULT_RULES = [
    ("E-TRANSFER", "Income"),
    ("WESCAM", "Income"),
    ("CANADA           PRO", "Income"),
    ("CANADA PRO", "Income"),
    ("GST", "Income"),
    ("NSLSC", "Income"),
    ("CDACARBON", "Income"),
    ("PAYMENT - THANK YOU", "Credit Payment"),
    ("TFR-TO", "Transfer"),
    ("TFR-FR", "Transfer"),
    ("TFR-", "Transfer"),
    ("TIM HORTONS", "Dining"),
    ("MCDONALDS", "Dining"),
    ("SUBWAY", "Dining"),
    ("STARBUCKS", "Dining"),
    ("PIZZA", "Dining"),
    ("TWITCH", "Subscriptions"),
    ("NETFLIX", "Subscriptions"),
    ("SPOTIFY", "Subscriptions"),
    ("FANATICS", "Shopping"),
    ("AMAZON", "Shopping"),
    ("WALMART", "Shopping"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")


def load_existing_rules():
    if os.path.exists(OUTPUT_FILE):
        try:
            df = pd.read_excel(OUTPUT_FILE, sheet_name="Category Rules")
            rules = [(str(row["Keyword"]).strip(), str(row["Category"]).strip())
                     for _, row in df.iterrows()
                     if str(row["Keyword"]).strip() and str(row["Category"]).strip()]
            if rules:
                return rules
        except Exception:
            pass
    if os.path.exists(RULES_BACKUP):
        try:
            with open(RULES_BACKUP, "r", encoding="utf-8") as f:
                rules = [tuple(r) for r in json.load(f)]
            if rules:
                print(f"  finances.xlsx not found — loaded rules from backup.")
                return rules
        except Exception:
            pass
    return DEFAULT_RULES


def save_rules_backup(rules):
    with open(RULES_BACKUP, "w", encoding="utf-8") as f:
        json.dump(rules, f, indent=2)


def parse_chequing_csv(path, account):
    df = pd.read_csv(path, header=None, names=["Date", "Description", "Debit", "Credit", "Balance"])
    df["Date"] = pd.to_datetime(df["Date"].str.strip('"'), format="%Y-%m-%d")
    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce").fillna(0)
    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce").fillna(0)
    df["Amount"] = df["Credit"] - df["Debit"]
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
    df["Account"] = account
    df["Description"] = df["Description"].str.strip('"').str.strip()
    return df[["Date", "Account", "Description", "Amount", "Balance"]]


def parse_visa_csv(path, account):
    df = pd.read_csv(path, header=None, names=["Date", "Description", "Charge", "Payment", "Balance"])
    df["Date"] = pd.to_datetime(df["Date"], format="%m/%d/%Y")
    df["Charge"] = pd.to_numeric(df["Charge"], errors="coerce").fillna(0)
    df["Payment"] = pd.to_numeric(df["Payment"], errors="coerce").fillna(0)
    df["Amount"] = df["Payment"] - df["Charge"]
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce").fillna(0)
    df["Account"] = account
    df["Description"] = df["Description"].str.strip()
    return df[["Date", "Account", "Description", "Amount", "Balance"]]


def load_all_transactions():
    frames = []
    for account_dir in os.listdir(STATEMENTS_DIR):
        full_dir = os.path.join(STATEMENTS_DIR, account_dir)
        if not os.path.isdir(full_dir):
            continue
        csvs = glob.glob(os.path.join(full_dir, "*.csv"))
        is_chequing = "Chequing" in account_dir
        for csv_path in csvs:
            try:
                if is_chequing:
                    df = parse_chequing_csv(csv_path, account_dir)
                else:
                    df = parse_visa_csv(csv_path, account_dir)
                frames.append(df)
            except Exception as e:
                print(f"  Warning: could not parse {csv_path}: {e}")
    if not frames:
        raise ValueError("No transactions loaded — check Statements folder.")
    combined = pd.concat(frames, ignore_index=True)
    combined = combined.drop_duplicates(subset=["Date", "Account", "Description", "Amount", "Balance"])
    combined = combined.sort_values("Date", ascending=False).reset_index(drop=True)
    return combined


def assign_categories(df, rules):
    def match(description):
        desc_upper = description.upper()
        for keyword, category in rules:
            if keyword.upper() in desc_upper:
                return category
        return "Uncategorized"
    df["Category"] = df["Description"].apply(match)
    return df


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def write_transactions_sheet(wb, df):
    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Account", "Description", "Amount", "Category", "Balance"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    amount_col = headers.index("Amount") + 1
    balance_col = headers.index("Balance") + 1

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.append([row.Date.date(), row.Account, row.Description, row.Amount, row.Category, row.Balance])
        ws.cell(i, amount_col).number_format = '#,##0.00'
        ws.cell(i, balance_col).number_format = '#,##0.00'
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            ws.cell(i, col).fill = fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    last_row = len(df) + 1
    amount_col_letter = get_column_letter(amount_col)
    ws.conditional_formatting.add(
        f"{amount_col_letter}2:{amount_col_letter}{last_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        f"{amount_col_letter}2:{amount_col_letter}{last_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
    )

    balance_col = headers.index("Balance") + 1
    balance_col_letter = get_column_letter(balance_col)
    ws.column_dimensions[balance_col_letter].hidden = True

    ws.freeze_panes = "A2"
    auto_width(ws)


def write_summary_sheet(wb, df):
    ws = wb.create_sheet("Monthly Summary")
    df2 = df.copy()
    df2["Month"] = df2["Date"].dt.to_period("M").astype(str)
    pivot = df2.pivot_table(index="Month", columns="Category", values="Amount", aggfunc="sum", fill_value=0)
    pivot = pivot.reset_index()
    pivot = pivot.sort_values("Month", ascending=False)

    headers = list(pivot.columns)
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        ws.append(list(row))
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            cell = ws.cell(i, col)
            cell.fill = fill
            if col > 1:
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "B2"
    auto_width(ws)


def write_rules_sheet(wb, rules):
    ws = wb.create_sheet("Category Rules")
    ws.append(["Keyword", "Category"])
    style_header_row(ws, 1, 2)
    for keyword, category in rules:
        ws.append([keyword, category])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20


def write_conflicts_sheet(wb, df, rules):
    ws = wb.create_sheet("Conflicts")
    headers = ["Description", "Matched Keywords", "Matched Categories"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    unique_descs = df["Description"].drop_duplicates().sort_values()

    conflicts = []
    for desc in unique_descs:
        desc_upper = desc.upper()
        matches = [(kw, cat) for kw, cat in rules if kw.upper() in desc_upper]
        unique_cats = list(dict.fromkeys(cat for _, cat in matches))
        if len(unique_cats) > 1:
            conflicts.append((
                desc,
                ", ".join(kw for kw, _ in matches),
                ", ".join(unique_cats),
            ))

    for row in conflicts:
        ws.append(list(row))

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 35

    if not conflicts:
        ws.cell(3, 1).value = "No conflicts found."
        ws.cell(3, 1).font = Font(italic=True, color="888888")


def write_uncategorized_sheet(wb, df):
    ws = wb.create_sheet("Uncategorized")
    ws.append(["Description"])
    style_header_row(ws, 1, 1)
    unique_descs = (
        df[df["Category"] == "Uncategorized"]["Description"]
        .drop_duplicates()
        .sort_values()
    )
    for desc in unique_descs:
        ws.append([desc])
    ws.column_dimensions["A"].width = 55


def main():
    print("Loading category rules...")
    rules = load_existing_rules()
    print(f"  {len(rules)} rules loaded.")

    print("Reading CSVs...")
    df = load_all_transactions()
    print(f"  {len(df)} transactions loaded.")

    df = assign_categories(df, rules)
    uncategorized = (df["Category"] == "Uncategorized").sum()
    print(f"  {uncategorized} uncategorized transactions.")

    print(f"Writing {OUTPUT_FILE}...")
    wb = Workbook()
    wb.remove(wb.active)
    write_transactions_sheet(wb, df)
    write_summary_sheet(wb, df)
    write_rules_sheet(wb, rules)
    write_conflicts_sheet(wb, df, rules)
    write_uncategorized_sheet(wb, df)
    wb.save(OUTPUT_FILE)
    save_rules_backup(rules)
    print("Done! Open finances.xlsx to view your transactions.")
    print("Edit the 'Category Rules' sheet and re-run to update categories.")


if __name__ == "__main__":
    main()
